import csv
from io import StringIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from app.application.bank_resolver import resolve_bank_code
from app.application.models import AnalysisData, NormalizedTransaction, TransactionRow
from app.application.ofx_writer import build_ofx_statement


class ExportArtifactService:
    def write_analysis_report_workbook(
        self,
        output_path: Path,
        *,
        report_rows: list[TransactionRow],
        snapshot: AnalysisData,
    ) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Transacoes"
        sheet.append(["date", "description", "amount", "category", "reconciliation_status"])
        for item in self._active_rows(report_rows):
            sheet.append([item.date, item.description, item.amount, item.category, item.reconciliation_status])
        self._format_sheet(sheet)
        self._add_conciliacao_sheet(workbook, snapshot)
        workbook.save(output_path)

    def write_convert_artifacts(
        self,
        output_dir: Path,
        *,
        report_rows: list[TransactionRow],
        ofx_account_type: str | None = None,
        layout_inference_name: str | None = None,
        opening_balance: float | None = None,
        closing_balance: float | None = None,
        bank_branch: str | None = None,
        account_number: str | None = None,
        bank_code: str | None = None,
    ) -> None:
        active_rows = self._active_rows(report_rows)
        normalized_transactions = [
            NormalizedTransaction(
                date=item.date,
                description=item.description,
                amount=item.amount,
                type="inflow" if item.amount >= 0 else "outflow",
            )
            for item in active_rows
        ]

        resolved_bank_code = resolve_bank_code(
            bank_code_override=bank_code,
            layout_inference_name=layout_inference_name,
        )
        (output_dir / "converted.ofx").write_text(
            build_ofx_statement(
                normalized_transactions,
                account_type=ofx_account_type,
                closing_balance=closing_balance,
                bank_branch=bank_branch,
                account_number=account_number,
                bank_id=resolved_bank_code,
            ),
            encoding="utf-8",
        )

        csv_buffer = StringIO()
        writer = csv.writer(csv_buffer)
        writer.writerow(["date", "description", "amount", "category", "reconciliation_status"])
        for item in active_rows:
            writer.writerow([item.date, item.description, item.amount, item.category, item.reconciliation_status])
        (output_dir / "converted.csv").write_text(csv_buffer.getvalue(), encoding="utf-8")

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Conversao"
        sheet.append(["Data", "Historico", "Credito", "Debito", "Saldo"])
        for item in active_rows:
            amount = float(item.amount)
            credit = round(amount, 2) if amount > 0 else None
            debit = round(abs(amount), 2) if amount < 0 else None
            sheet.append([self._format_convert_date(item.date), item.description, credit, debit, item.running_balance])
        self._format_sheet(sheet)

        opening_balance_value, closing_balance_value = self._resolve_balance_bounds(
            active_rows,
            opening_balance_hint=opening_balance,
            closing_balance_hint=closing_balance,
        )
        summary_sheet = workbook.create_sheet(title="Resumo")
        summary_sheet.append(["Campo", "Valor"])
        summary_sheet.append(["Saldo anterior", opening_balance_value])
        summary_sheet.append(["Saldo final", closing_balance_value])
        self._format_sheet(summary_sheet)
        workbook.save(output_dir / "converted.xlsx")

    def _active_rows(self, rows: list[TransactionRow]) -> list[TransactionRow]:
        return [item for item in rows if not item.is_deleted]

    def _format_sheet(self, sheet) -> None:
        if sheet.max_row >= 1 and sheet.max_column >= 1:
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions

        for column_index in range(1, sheet.max_column + 1):
            column_letter = get_column_letter(column_index)
            max_len = 0
            for row_index in range(1, sheet.max_row + 1):
                cell_value = sheet.cell(row=row_index, column=column_index).value
                max_len = max(max_len, len(str(cell_value)) if cell_value is not None else 0)
            sheet.column_dimensions[column_letter].width = min(max(max_len + 2, 12), 80)

    def _add_conciliacao_sheet(self, workbook: Workbook, data: AnalysisData) -> None:
        sheet = workbook.create_sheet(title="Conciliacao")
        sheet.append(["metric", "value"])
        sheet.append(["matched_groups", data.matched_groups])
        sheet.append(["reversed_entries", data.reversed_entries])
        sheet.append(["potential_duplicates", data.potential_duplicates])
        sheet.append([])
        sheet.append(["date", "description", "amount", "category", "reconciliation_status"])
        for item in data.preview_transactions:
            if item.reconciliation_status != "unmatched":
                sheet.append([item.date, item.description, item.amount, item.category, item.reconciliation_status])

        if sheet.max_row == 6:
            sheet.append(["-", "No reconciled entries in preview", "", "", "unmatched"])

        self._format_sheet(sheet)

    def _format_convert_date(self, value: str) -> str:
        raw = str(value or "").strip()
        if len(raw) == 10 and raw[4] == "-" and raw[7] == "-":
            year, month, day = raw.split("-")
            if year.isdigit() and month.isdigit() and day.isdigit():
                return f"{day}-{month}-{year}"
        return raw

    def _resolve_balance_bounds(
        self,
        rows: list[TransactionRow],
        *,
        opening_balance_hint: float | str | None = None,
        closing_balance_hint: float | str | None = None,
    ) -> tuple[float | None, float | None]:
        opening_balance = self._coerce_balance_value(opening_balance_hint)
        if opening_balance is None:
            for item in rows:
                if self._is_opening_balance_row(item):
                    opening_balance = round(float(item.amount), 2)
                    break
            if opening_balance is None:
                for item in rows:
                    if item.running_balance is None:
                        continue
                    opening_balance = round(float(item.running_balance) - float(item.amount), 2)
                    break

        manual_closing_balance = self._coerce_balance_value(closing_balance_hint)
        if manual_closing_balance is not None:
            return opening_balance, manual_closing_balance

        last_balance_index: int | None = None
        closing_balance: float | None = None
        for index, item in enumerate(rows):
            if item.running_balance is None:
                continue
            last_balance_index = index
            closing_balance = round(float(item.running_balance), 2)
        if closing_balance is not None and last_balance_index is not None:
            trailing_amount = sum(float(item.amount) for item in rows[last_balance_index + 1 :])
            return opening_balance, round(closing_balance + trailing_amount, 2)

        if opening_balance is not None:
            transaction_amounts = sum(float(item.amount) for item in rows if not self._is_opening_balance_row(item))
            return opening_balance, round(opening_balance + transaction_amounts, 2)

        return None, None

    def _coerce_balance_value(self, value: float | str | None) -> float | None:
        if value is None or value == "":
            return None
        try:
            return round(float(value), 2)
        except (TypeError, ValueError):
            return None

    def _is_opening_balance_row(self, row: TransactionRow) -> bool:
        normalized = " ".join(str(row.description or "").strip().upper().split())
        return normalized in {"SALDO ANTERIOR", "SALDO INICIAL"}
