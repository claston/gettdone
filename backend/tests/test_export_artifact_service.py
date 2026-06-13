from io import BytesIO

from openpyxl import load_workbook

from app.application.export_artifact_service import ExportArtifactService
from app.application.models import AnalysisData, TransactionRow


def _build_rows() -> list[TransactionRow]:
    return [
        TransactionRow(
            date="2026-04-01",
            description="SALARIO",
            amount=100.0,
            category="Outros",
            reconciliation_status="unmatched",
            running_balance=150.0,
        ),
        TransactionRow(
            date="2026-04-02",
            description="TARIFA",
            amount=-20.0,
            category="Outros",
            reconciliation_status="matched",
            running_balance=130.0,
        ),
    ]


def _build_snapshot() -> AnalysisData:
    rows = _build_rows()
    return AnalysisData(
        analysis_id="an_export123",
        file_type="pdf",
        upload_filename="extrato.pdf",
        transactions_total=2,
        total_inflows=100.0,
        total_outflows=-20.0,
        net_total=80.0,
        preview_transactions=rows,
        report_transactions=rows,
        matched_groups=1,
        reversed_entries=0,
        potential_duplicates=0,
    )


def test_write_convert_artifacts_creates_ofx_csv_and_xlsx(tmp_path) -> None:
    service = ExportArtifactService()

    service.write_convert_artifacts(
        tmp_path,
        report_rows=_build_rows(),
        layout_inference_name="bradesco_net_empresa_extrato_mensal_por_periodo_v1",
        opening_balance=50.0,
        closing_balance=130.0,
        bank_branch="3456-7",
        account_number="12345-6",
    )

    ofx_text = (tmp_path / "converted.ofx").read_text(encoding="utf-8")
    csv_text = (tmp_path / "converted.csv").read_text(encoding="utf-8")
    xlsx_path = tmp_path / "converted.xlsx"

    assert "<STMTTRN>" in ofx_text
    assert "<BRANCHID>34567" in ofx_text
    assert "<ACCTID>123456" in ofx_text
    assert "<BANKID>237" in ofx_text
    assert "2026-04-01,SALARIO,100.0,Outros,unmatched" in csv_text

    workbook = load_workbook(filename=xlsx_path, data_only=True)
    sheet = workbook["Conversao"]
    summary = workbook["Resumo"]
    assert sheet.cell(row=2, column=1).value == "01-04-2026"
    assert sheet.cell(row=2, column=3).value == 100
    assert sheet.cell(row=3, column=4).value == 20
    assert summary.cell(row=2, column=1).value == "Saldo anterior"
    assert summary.cell(row=2, column=2).value == 50
    assert summary.cell(row=3, column=1).value == "Saldo final"
    assert summary.cell(row=3, column=2).value == 130


def test_write_analysis_report_workbook_creates_transacoes_and_conciliacao(tmp_path) -> None:
    service = ExportArtifactService()
    output_path = tmp_path / "report.xlsx"

    service.write_analysis_report_workbook(
        output_path,
        report_rows=_build_rows(),
        snapshot=_build_snapshot(),
    )

    workbook = load_workbook(filename=BytesIO(output_path.read_bytes()), data_only=True)
    transacoes = workbook["Transacoes"]
    conciliacao = workbook["Conciliacao"]

    assert transacoes.cell(row=1, column=1).value == "date"
    assert transacoes.cell(row=2, column=2).value == "SALARIO"
    assert conciliacao.cell(row=2, column=1).value == "matched_groups"
    assert conciliacao.cell(row=2, column=2).value == 1
    assert conciliacao.cell(row=7, column=2).value == "TARIFA"
