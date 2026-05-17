from io import BytesIO
from pathlib import Path

from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.application import AccessControlService, ReportService, TempAnalysisStorage
from app.application.models import AnalysisData, TransactionRow
from app.dependencies import get_access_control_service, get_report_service
from app.main import app


def _build_analysis_data(analysis_id: str = "an_convert123") -> AnalysisData:
    return AnalysisData(
        analysis_id=analysis_id,
        file_type="pdf",
        upload_filename="extrato_nubank.pdf",
        layout_inference_name="bradesco_net_empresa_extrato_mensal_por_periodo_v1",
        transactions_total=1,
        total_inflows=100.0,
        total_outflows=-20.0,
        net_total=80.0,
        preview_transactions=[
            TransactionRow(
                date="2026-04-01",
                description="TEST",
                amount=-20.0,
                category="Outros",
                reconciliation_status="unmatched",
            )
        ],
        report_transactions=[
            TransactionRow(
                date="2026-04-01",
                description="TEST",
                amount=-20.0,
                category="Outros",
                reconciliation_status="unmatched",
            )
        ],
    )


def build_client(tmp_path: Path) -> TestClient:
    storage = TempAnalysisStorage(root_dir=tmp_path, ttl_seconds=3600)
    storage.save_analysis(_build_analysis_data())
    access_control = AccessControlService(
        state_file=tmp_path / "access-control-state.json",
        token_secret="test-secret",
    )
    owner = access_control.resolve_identity(anonymous_fingerprint="fp-owner", user_token=None)
    storage.set_convert_owner(
        analysis_id="an_convert123",
        identity_type=owner.identity_type,
        identity_id=owner.identity_id,
    )
    app.dependency_overrides[get_report_service] = lambda: ReportService(storage=storage)
    app.dependency_overrides[get_access_control_service] = lambda: access_control
    return TestClient(app)


def build_client_with_user_owner(tmp_path: Path) -> tuple[TestClient, str]:
    storage = TempAnalysisStorage(root_dir=tmp_path, ttl_seconds=3600)
    storage.save_analysis(_build_analysis_data("an_user_owned"))
    access_control = AccessControlService(
        state_file=tmp_path / "access-control-state.json",
        token_secret="test-secret",
    )
    user = access_control.register_user(name="Erica", email="erica@example.com", password="strong-pass")
    storage.set_convert_owner(
        analysis_id="an_user_owned",
        identity_type="user",
        identity_id=user.user_id,
    )
    app.dependency_overrides[get_report_service] = lambda: ReportService(storage=storage)
    app.dependency_overrides[get_access_control_service] = lambda: access_control
    return TestClient(app), user.token


def test_convert_report_download_happy_path(tmp_path: Path) -> None:
    client = build_client(tmp_path)

    response = client.get("/convert-report/an_convert123?format=ofx&anonymous_fingerprint=fp-owner")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/x-ofx")
    assert "extrato_nubank.ofx" in response.headers["content-disposition"]
    assert "<STMTTRN>" in response.text
    app.dependency_overrides.clear()


def test_convert_report_download_csv_happy_path(tmp_path: Path) -> None:
    client = build_client(tmp_path)

    response = client.get("/convert-report/an_convert123?format=csv&anonymous_fingerprint=fp-owner")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("text/csv")
    assert "extrato_nubank.csv" in response.headers["content-disposition"]
    assert "date,description,amount" in response.text
    app.dependency_overrides.clear()


def test_convert_report_download_xlsx_matches_conversion_review_layout(tmp_path: Path) -> None:
    client = build_client(tmp_path)

    response = client.get("/convert-report/an_convert123?format=xlsx&anonymous_fingerprint=fp-owner")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    assert "extrato_nubank.xlsx" in response.headers["content-disposition"]

    workbook = load_workbook(filename=BytesIO(response.content), data_only=True)
    sheet = workbook.active
    headers = [sheet.cell(row=1, column=idx).value for idx in range(1, 5)]
    assert headers == ["Data", "Historico", "Credito", "Debito"]
    assert sheet.cell(row=2, column=1).value == "01-04-2026"
    assert sheet.cell(row=2, column=2).value == "TEST"
    assert sheet.cell(row=2, column=3).value is None
    assert sheet.cell(row=2, column=4).value == 20
    app.dependency_overrides.clear()


def test_convert_report_returns_not_found_for_missing_analysis(tmp_path: Path) -> None:
    client = build_client(tmp_path)

    response = client.get("/convert-report/an_missing?format=ofx&anonymous_fingerprint=fp-owner")

    assert response.status_code == 404
    assert response.json()["detail"] == "Analysis not found"
    app.dependency_overrides.clear()


def test_convert_report_rejects_access_from_other_identity(tmp_path: Path) -> None:
    client = build_client(tmp_path)

    response = client.get("/convert-report/an_convert123?format=ofx&anonymous_fingerprint=fp-other")

    assert response.status_code == 403
    assert response.json()["detail"] == "Access denied for this analysis."
    app.dependency_overrides.clear()


def test_convert_report_accepts_bearer_user_token(tmp_path: Path) -> None:
    client, user_token = build_client_with_user_owner(tmp_path)

    response = client.get(
        "/convert-report/an_user_owned?format=ofx",
        headers={"authorization": f"Bearer {user_token}"},
    )

    assert response.status_code == 200
    assert response.headers["content-type"].startswith("application/x-ofx")
    assert "<STMTTRN>" in response.text
    app.dependency_overrides.clear()


def test_convert_report_ofx_accepts_closing_balance_override(tmp_path: Path) -> None:
    client = build_client(tmp_path)

    response = client.get(
        "/convert-report/an_convert123?format=ofx&closing_balance=56276.06&anonymous_fingerprint=fp-owner"
    )

    assert response.status_code == 200
    assert "<LEDGERBAL>" in response.text
    assert "<BALAMT>56276.06" in response.text
    assert "<BRANCHID>0001" in response.text
    assert "<ACCTID>000000" in response.text
    app.dependency_overrides.clear()


def test_convert_report_ofx_accepts_bank_branch_and_account_number_override(tmp_path: Path) -> None:
    client = build_client(tmp_path)

    response = client.get(
        "/convert-report/an_convert123?format=ofx&bank_branch=3456-7&account_number=12345-6&anonymous_fingerprint=fp-owner"
    )

    assert response.status_code == 200
    assert "<BRANCHID>34567" in response.text
    assert "<ACCTID>123456" in response.text
    app.dependency_overrides.clear()


def test_convert_report_ofx_infers_bank_code_from_layout(tmp_path: Path) -> None:
    client = build_client(tmp_path)

    response = client.get("/convert-report/an_convert123?format=ofx&anonymous_fingerprint=fp-owner")

    assert response.status_code == 200
    assert "<BANKID>237" in response.text
    app.dependency_overrides.clear()


def test_convert_report_ofx_accepts_bank_code_override(tmp_path: Path) -> None:
    client = build_client(tmp_path)

    response = client.get("/convert-report/an_convert123?format=ofx&bank_code=001&anonymous_fingerprint=fp-owner")

    assert response.status_code == 200
    assert "<BANKID>001" in response.text
    app.dependency_overrides.clear()
